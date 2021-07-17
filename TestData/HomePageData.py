import openpyxl

class Homepagedata:

    #test_homepage_data =

    @staticmethod
    def get_testdata():


        book = openpyxl.load_workbook("C:\\SelPyPrj\\A_Project_1one\\TestData\\Excel Demo.xlsx")
        sheet = book.active
        data = []
        for i in range(2, sheet.max_row +1):
            Dict = {}
            for j in range(2, sheet.max_column+1):
                Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
            data.append(Dict)
            del Dict

        return data